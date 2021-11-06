# -*- coding: utf-8 -*-
# @公众号 :Web图书馆 代码详解：http://t.cn/A6Ihdrxf
# @Software: PyCharm 安装教程：https://mp.weixin.qq.com/s/a0zoCo9DacvdpIoz1LEN3Q
# @Description:
# Python全套学习资源：https://mp.weixin.qq.com/s/G_5cY05Qoc_yCXGQs4vIeg

import openpyxl

district_list = ['上海', '山东', '广东', '广西', '江苏', '河南', '浙江']
zxl_总销量 = 0
zth_总退货 = 0
thl_退货率 = 0
zyye_总营业额 = 0
zyf_总运费 = 0
zly_总利润 = 0
for district in district_list:

    workbook = openpyxl.load_workbook('Excel案例1/{}.xlsx'.format(district))
    sheet = workbook['语文']
    max_row = sheet.max_row
    print(max_row)
    for row in range(2, max_row + 1):
        xl_销量 = sheet['D' + str(row)].value
        zxl_总销量 += xl_销量
        th_退货 = sheet['E' + str(row)].value
        zth_总退货 += th_退货

        yye_营业额 = (xl_销量 - th_退货) * sheet['F' + str(row)].value
        zyye_总营业额 += yye_营业额

        yf_运费 = sheet['G' + str(row)].value
        zyf_总运费 += yf_运费

print('总销量 = {}'.format(zxl_总销量) )
print('zth_总退货 = {}'.format(zth_总退货))
thl_退货率 = "%.2f%%" % (zth_总退货 / zxl_总销量 * 100)
print('thl_退货率 = {}'.format(thl_退货率))
print('zyye_总营业额 = {}'.format(zyye_总营业额))
print('zyf_总运费 = {}'.format(zyf_总运费))
zly_总利润 = zyye_总营业额-zyf_总运费
print('zly_总利润 = {}'.format(zly_总利润))

workbook3 = openpyxl.load_workbook('res.xlsx')
sheet3 = workbook3['Sheet3']
sheet3['B2'] = zxl_总销量
sheet3['C2'] = zth_总退货
sheet3['D2'] = thl_退货率
sheet3['E2'] = zyye_总营业额
sheet3['F2'] = zyf_总运费
sheet3['G2'] = zly_总利润
workbook3.save('res.xlsx')

